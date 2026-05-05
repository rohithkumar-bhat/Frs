"""
Extract attendance data from a Google Sheets spreadsheet.

Downloads the spreadsheet as XLSX to support multiple tabs,
parses month-wise and daily-wise attendance blocks,
merges per-employee, and writes static/data.js for the dashboard.
"""

import json
import os
import re
from datetime import datetime

import requests
from dotenv import load_dotenv


def get_sheet_url_xlsx(url_var="GOOGLE_SHEET_URL"):
    """Read the Google Sheets URL from .env and convert to XLSX export URL."""
    load_dotenv(override=True)
    raw_url = os.getenv(url_var, "")
    if not raw_url:
        return None

    # Extract the spreadsheet ID
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', raw_url)
    if not match:
        raise ValueError(f"Could not extract spreadsheet ID from {url_var}: {raw_url}")

    sheet_id = match.group(1)
    import time
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx&t={int(time.time())}"



def parse_date_header(cell_val):
    """
    Convert date header from various formats to YYYY-MM-DD.
    Returns None if the header is not a date.
    """
    if cell_val is None:
        return None
    
    if isinstance(cell_val, datetime):
        return cell_val.strftime('%Y-%m-%d')
    
    date_str = str(cell_val).strip()
    if not date_str:
        return None
        
    # Handle full datetime strings from str(datetime)
    if ' ' in date_str:
        date_str = date_str.split(' ')[0]

    # Try various formats
    for fmt in ('%m/%d/%Y', '%m-%d-%Y', '%Y-%m-%d', '%d-%m-%Y', '%d-%b-%Y', '%d-%B-%Y'):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def parse_time_val(val):
    """Convert datetime.time, float (Excel time), or string time to HH:MM format."""
    if val is None or val == 0 or val == 0.0 or str(val).strip() == '------' or str(val).strip() == '':
        return None
        
    if isinstance(val, (datetime,)):
        return val.strftime('%H:%M')
    
    # Handle datetime.time
    import datetime as dt
    if isinstance(val, dt.time):
        return val.strftime('%H:%M')
    
    # Handle Excel float time (fraction of a day)
    if isinstance(val, (int, float)):
        total_seconds = int(val * 86400)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"

    # Try parsing string if it's "HH:MM AM/PM" or just "HH:MM"
    s = str(val).strip()
    for fmt in ('%I:%M %p', '%I:%M%p', '%H:%M:%S', '%H:%M'):
        try:
            return dt.datetime.strptime(s, fmt).strftime('%H:%M')
        except ValueError:
            continue
    return s



EXCLUDED_NAMES = [
    "Chennupati Murali",
    "Gade Ramya Manasa",
    "Gogusetty Avinash",
    "Komal Agarwal"
]

def apply_overrides(employees):
    """Apply manual attendance overrides and filter out excluded employees."""
    # 1. Filter out excluded names
    filtered_employees = [e for e in employees if e.get('Employee Name') not in EXCLUDED_NAMES]
    
    # 2. Apply existing overrides
    emp_lookup = {str(emp['Employee ID']): emp for emp in filtered_employees}


    overrides = {
        "EL1709": [  # Jyothi Babu Reddy
            "2026-02-23", "2026-02-26", "2026-02-28",
            "2026-03-11", "2026-03-17"
        ],
        "EL170204": [  # Punna Reddy
            "2026-02-17", "2026-02-27",
            "2026-03-10", "2026-03-13"
        ],
        "EL220239": [  # Vasundhara Reddy
            "2026-02-16", "2026-02-25",
            "2026-03-13", "2026-03-21"
        ]
    }

    for emp_id, na_dates in overrides.items():
        emp = emp_lookup.get(emp_id)
        if emp:
            for d in na_dates:
                emp[d] = "NA"
    return filtered_employees



def run_extraction():
    """Main entry point to fetch and save the data."""
    try:
        import openpyxl
        import time

        
        # 1. Get the export URL (XLSX)
        export_url = get_sheet_url_xlsx("GOOGLE_SHEET_URL")
        if not export_url:
            raise ValueError("No 'GOOGLE_SHEET_URL' found in .env or environment")
        print(f"Fetching XLSX from Google Sheets...")

        # 2. Fetch XLSX data
        response = requests.get(export_url, timeout=60)
        response.raise_for_status()
        
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        # 3. Identify sheets
        month_sheet = None
        daily_sheets = []
        print(f"Workbook sheets found: {', '.join(wb.sheetnames)}")
        for name in wb.sheetnames:
            lname = name.lower().strip()
            if 'month' in lname: 
                month_sheet = wb[name]
                print(f"  Selected for Monthly parsing: '{name}'")
            elif 'daily' in lname: 
                daily_sheets.append(wb[name])
                print(f"  Selected for Daily parsing: '{name}'")
        
        if not month_sheet:
            month_sheet = wb.active
            print(f"  No 'Month' sheet found. Falling back to active sheet: '{month_sheet.title}'")

        # Check for a dedicated DAILY_SHEET_URL
        daily_export_url = get_sheet_url_xlsx("DAILY_SHEET_URL")
        if daily_export_url:
            print(f"Fetching Daily XLSX from dedicated Google Sheet URL...")
            response_daily = requests.get(daily_export_url, timeout=60)
            response_daily.raise_for_status()
            
            wb_daily = openpyxl.load_workbook(BytesIO(response_daily.content), data_only=True)
            print(f"Dedicated Daily Workbook sheets: {', '.join(wb_daily.sheetnames)}")
            for name in wb_daily.sheetnames:
                daily_sheets.append(wb_daily[name])
                print(f"  Added daily sheet: '{name}'")



        # 4. Parse Monthly Data
        rows = list(month_sheet.iter_rows(values_only=True))
        employee_map = {}
        months_found = []

        i = 0
        while i < len(rows):
            row = rows[i]
            sr_no_idx = None
            for col_idx, cell in enumerate(row):
                if cell and str(cell).strip() == 'Sr.No.':
                    sr_no_idx = col_idx
                    break

            if sr_no_idx is not None:
                # Header row found
                headers = []
                for col_idx, cell in enumerate(row):
                    parsed_date = parse_date_header(cell)
                    if parsed_date:
                        headers.append(parsed_date)
                    else:
                        headers.append(str(cell).strip() if cell is not None else None)

                date_keys_in_block = [h for h in headers if h and re.match(r'^\d{4}-\d{2}-\d{2}$', h)]
                month_str = date_keys_in_block[0][:7] if date_keys_in_block else "unknown"
                if month_str not in months_found and month_str != "unknown":
                    months_found.append(month_str)
                    print(f"  Found month block: {month_str}")

                i += 1
                while i < len(rows):
                    data_row = rows[i]
                    if not data_row or len(data_row) <= sr_no_idx or data_row[sr_no_idx] is None:
                        i += 1
                        continue
                    
                    sr_val = str(data_row[sr_no_idx]).strip()
                    if sr_val.startswith('NA-') or sr_val == '':
                        # Check some other cells to see if it's really the end
                        is_end = True
                        if len(data_row) > sr_no_idx + 1 and data_row[sr_no_idx+1]: is_end = False
                        if is_end:
                            i += 1
                            break

                    emp_data = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(data_row) and header:
                            val = data_row[col_idx]
                            target_key = header
                            if header == 'Attendence' or header == 'Percentage' or 'Total Days Attended' in header or 'Attendance' in header:
                                target_key = f"{month_str}-Attendence" if ('Attendence' in header or 'Total' in header or 'Attendance' in header) else f"{month_str}-{header}"
                            
                            if isinstance(val, (datetime,)):
                                val = val.strftime('%H:%M' if val.hour or val.minute else '%Y-%m-%d')
                            elif hasattr(val, 'strftime'): # time
                                val = val.strftime('%H:%M')
                            
                            emp_data[target_key] = val

                    emp_id = str(emp_data.get('Employee ID', '')).strip()
                    if emp_id and not emp_id.startswith('NA'):
                        if emp_id not in employee_map:
                            employee_map[emp_id] = {
                                'Sr.No.': emp_data.get('Sr.No.'),
                                'Employee Name': emp_data.get('Employee Name'),
                                'Employee ID': emp_id,
                                'Branch': emp_data.get('Branch'),
                                'daily_details': {} 
                            }
                        for key, val in emp_data.items():
                            if key not in ('Sr.No.', 'Employee Name', 'Employee ID', 'Branch'):
                                # Avoid overwriting newer month bits if multiple blocks exist
                                employee_map[emp_id][key] = val
                    i += 1
            else:
                i += 1

        # 5. Parse Daily Details if sheets exist
        for daily_sheet in daily_sheets:
            print(f"  Parsing daily details from '{daily_sheet.title}'")
            daily_rows = list(daily_sheet.iter_rows(values_only=True))
            
            current_date = None
            col_map = None

            for row in daily_rows:
                if not row or all(c is None for c in row):
                    continue

                row_str_vals = [str(c).strip() if c is not None else '' for c in row]

                # 1. Date Detection in the row
                for cell in row:
                    if cell is None: continue
                    cell_str = str(cell).strip()
                    # Support: 04-May-2026, 04-05-2026, 4/5/2026, 2026-05-04
                    date_match = re.search(r'(\d{1,4}[-/]\d{1,2}[-/]\d{1,4})|(\d{1,2}-[A-Za-z]{3,}-\d{4})', cell_str)
                    if date_match:
                        parsed = parse_date_header(date_match.group(0))
                        if parsed:
                            current_date = parsed
                            print(f"    Detected date in sheet: {current_date} from '{cell_str}'")
                            break

                
                # 2. Header Detection
                if any(h in row_str_vals for h in ['Employee ID', 'ID', 'Check In Time', 'Login']):
                    def find_col(aliases):
                        for i, h in enumerate(row_str_vals):
                            h_low = h.lower()
                            if any(a.lower() in h_low for a in aliases):
                                return i
                        return -1

                    col_map = {
                        'ID': find_col(['Employee ID', 'ID']),
                        'Login': find_col(['Check In Time', 'In Time', 'Login', 'Check-In', 'Time In']),
                        'Logout': find_col(['Check Out Time', 'Out Time', 'Logout', 'Check-Out', 'Time Out']),
                        'Total': find_col(['Total Duration', 'Duration', 'Working Hours', 'Total Hours', 'Work Time']),
                        'Break': find_col(['Break Time', 'Break Duration', 'Break']),
                        'Date': find_col(['Date', 'Day']),
                    }
                    print(f"    Detected daily columns: {col_map}")
                    continue


                # 3. Data Processing
                if col_map and col_map['ID'] != -1 and col_map['ID'] < len(row):
                    emp_id = str(row[col_map['ID']]).strip()
                    if not emp_id or emp_id.lower() == 'none' or emp_id == 'Employee ID':
                        continue
                    
                    # Row-specific date override
                    row_date = current_date
                    if col_map['Date'] != -1 and col_map['Date'] < len(row):
                        parsed_row_date = parse_date_header(row[col_map['Date']])
                        if parsed_row_date:
                            row_date = parsed_row_date

                    if emp_id in employee_map and row_date:
                        login = parse_time_val(row[col_map['Login']]) if col_map['Login'] != -1 and col_map['Login'] < len(row) else None
                        logout = parse_time_val(row[col_map['Logout']]) if col_map['Logout'] != -1 and col_map['Logout'] < len(row) else None
                        total = parse_time_val(row[col_map['Total']]) if col_map['Total'] != -1 and col_map['Total'] < len(row) else None
                        break_t = parse_time_val(row[col_map['Break']]) if col_map['Break'] != -1 and col_map['Break'] < len(row) else None
                        
                        # Store the granular details for the modal
                        employee_map[emp_id]['daily_details'][row_date] = {
                            'login': login,
                            'logout': logout,
                            'break': break_t,
                            'total': total
                        }
                        print(f"      Mapped: {emp_id} on {row_date} (In: {login}, Out: {logout}, Work: {total})")
                        
                        # Update the calendar status
                        # Prefer 'total' hours if it looks like a time duration, otherwise fallback to 'Present'
                        current_status = str(employee_map[emp_id].get(row_date, '')).strip().upper()
                        if not current_status or current_status in ('NA', '--', 'NONE', '0', ''):
                            if total:
                                employee_map[emp_id][row_date] = total
                            elif login:
                                employee_map[emp_id][row_date] = login
                            else:
                                employee_map[emp_id][row_date] = 'Present'

                    elif row_date:
                        print(f"      Warning: ID '{emp_id}' from daily sheet not found in monthly list.")



        # 6. Final Pass: Recalculate Attendance Totals and Percentages
        # This ensures that auto-populated daily data updates the summary counts
        for emp_id, emp in employee_map.items():
            # Identify all months present for this employee
            months = set()
            for key in emp.keys():
                if len(key) == 7 and key[4] == '-': # YYYY-MM
                    months.add(key)
            
            for month_prefix in months:
                days_present = 0
                total_working_days = 0
                
                # Iterate through all days of that month
                for key, val in emp.items():
                    if key.startswith(month_prefix) and len(key) == 10: # YYYY-MM-DD
                        v_str = str(val or '').strip().upper()
                        # Skip non-working days
                        if v_str in ('SUNDAY', 'HOLIDAY') or 'HOLIDAY' in v_str:
                            continue
                        
                        total_working_days += 1
                        
                        # Count as present if it has a time format or is 'PRESENT'
                        is_present = False
                        if v_str == 'PRESENT':
                            is_present = True
                        elif ':' in v_str: # HH:MM format
                            is_present = True
                        elif v_str and v_str not in ('LEAVE', 'NA', '--', 'NONE', '0'):
                            is_present = True
                            
                        if is_present:
                            days_present += 1
                
                # Update the summary keys
                if total_working_days > 0:
                    emp[f"{month_prefix}-Attendence"] = days_present
                    emp[f"{month_prefix}-Percentage"] = round((days_present / total_working_days) * 100, 2)

        # 7. Finalize
        employees = apply_overrides(list(employee_map.values()))

        output_path = os.path.join("static", "data.js")

        last_updated = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with open(output_path, "w") as f:
            f.write(f"// Generated on {last_updated}\n")
            f.write(f"const LAST_UPDATED = \"{last_updated}\";\n")
            f.write("const attendanceData = ")
            json.dump(employees, f, indent=2)
            f.write(";")

        print(f"SUCCESS: data.js created with {len(employees)} employees and {len(months_found)} months: {', '.join(months_found)}")
        return employees

    except Exception as e:
        import traceback
        error_msg = f"Extraction failed: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        raise e

def main():

    run_extraction()

if __name__ == "__main__":
    main()

