import openpyxl

def peek_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True, max_row=10):
                print(row)
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    peek_excel("Book1.xlsx")
